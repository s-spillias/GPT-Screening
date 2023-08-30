recommender = None
import os
import openpyxl
import pandas as pd

# Read PDF

def download_pdf(url, output_path):
    urllib.request.urlretrieve(url, output_path)

def preprocess(text):
    text = text.replace('\n', ' ')
    text = re.sub('\s+', ' ', text)
    return text

def split_text_at_acknowledgements(text):
    target_word_pattern = r'a\s*c\s*k\s*n\s*o\s*w\s*l\s*e\s*d\s*g\s*(e\s*m\s*e\s*n\s*t\s|m\s*e\s*n\s*t\s*)'
    match = re.search(target_word_pattern, text, re.I)

    if match:
        index = match.start()
        part1 = text[:index]
        part2 = text[index:]

        parts = [part1,part2]
    else:
        # Handle case where the target words are not found
        parts = [text]
    return parts

def pdf_to_text(path, start_page=1, end_page=None):
    doc = fitz.open(path)
    total_pages = doc.page_count

    if end_page is None:
        end_page = total_pages

    text_list = []
    frontmatter = []
    backmatter = []
    in_frontmatter = True
    in_backmatter = False
    for i in range(start_page-1, end_page):
        text = doc.load_page(i).get_text("text")
        text = preprocess(text)
        if in_backmatter:
            backmatter.append(text)
            continue
        if in_frontmatter:
            split_at_abstract = text.split('Abstract')
            if (len(split_at_abstract)>1):
                in_frontmatter = False
                text = split_at_abstract[1]
                frontmatter.append(split_at_abstract[0])
            else:
                frontmatter.append(text)
        split_at_acknowledge = split_text_at_acknowledgements(text)
        if (len(split_at_acknowledge)>1):
            in_backmatter = True
            backmatter.append(split_at_acknowledge[1])
            text = split_at_acknowledge[0]
        text_list.append(text)
    doc.close()
    return text_list,frontmatter,backmatter


def text_to_chunks(text):
    text_splitter = RecursiveCharacterTextSplitter(
        #separator="\n",
        chunk_size=chunk_size,
        chunk_overlap=round(chunk_size/5),
       # length_function=len
    )
    chunks = text_splitter.split_text(text)
    return chunks

# Create Embeddings for PDF text.

import numpy as np
from sklearn.neighbors import NearestNeighbors
from transformers import AutoTokenizer, AutoModel
import torch

class QASemanticSearch:
    
    def __init__(self):
        self.tokenizer = AutoTokenizer.from_pretrained("sentence-transformers/stsb-roberta-large")
        self.model = AutoModel.from_pretrained("sentence-transformers/stsb-roberta-large")
        self.fitted = False
    
    def fit(self, chunks, batch=1000, n_neighbors=5):
        self.chunks = chunks
        embeddings = self.get_text_embedding(chunks, batch=batch)
        
        self.nn = NearestNeighbors(n_neighbors=n_neighbors)
        self.nn.fit(embeddings)
        
        self.fitted = True
    
    def __call__(self, query, return_data=True):
        query_emb = self.get_text_embedding([query])
        neighbors = self.nn.kneighbors(query_emb, return_distance=False)[0]
        
        if return_data:
            return [self.chunks[i] for i in neighbors]
        else:
            return neighbors
    
    def get_text_embedding(self, texts, batch=1000):
        embeddings = []
        for i in range(0, len(texts), batch):
            text_batch = texts[i:(i+batch)]
            inputs = self.tokenizer(text_batch, return_tensors="pt", padding=True, truncation=True)
            with torch.no_grad():
                model_outputs = self.model(**inputs)
            emb_batch = model_outputs.last_hidden_state.mean(dim=1).numpy()
            embeddings.append(emb_batch)
        embeddings = np.vstack(embeddings)
        return embeddings


# Create Embeddings for PDF text.
class SemanticSearch:
    
    def __init__(self):
        self.tokenizer = AutoTokenizer.from_pretrained(embedding_model)
        self.model = AutoModel.from_pretrained(embedding_model)
        self.fitted = False
    
    def fit(self, data, batch=1000, n_neighbors=5):
        self.data = data
        self.embeddings = self.get_text_embedding(data, batch=batch)
        n_neighbors = min(n_neighbors, len(self.embeddings))
        self.nn = NearestNeighbors(n_neighbors=n_neighbors)
        self.nn.fit(self.embeddings)
        self.fitted = True
    
    def __call__(self, text, return_data=True):
        inp_emb = self.get_text_embedding([text])
        neighbors = self.nn.kneighbors(inp_emb, return_distance=False)[0]
        
        if return_data:
            return [self.data[i] for i in neighbors]
        else:
            return neighbors
    
    def get_text_embedding(self, texts, batch=1000):
        embeddings = []
        for i in range(0, len(texts), batch):
            text_batch = texts[i:(i+batch)]
            inputs = self.tokenizer(text_batch, return_tensors="pt", padding=True, truncation=True)
            with torch.no_grad():
                model_outputs = self.model(**inputs)
            emb_batch = model_outputs.last_hidden_state.mean(dim=1).numpy()
            embeddings.append(emb_batch)
        embeddings = np.vstack(embeddings)
        return embeddings

text_type = {'main':0,
             'frontmatter':1,
             'backmatter':2}

def load_recommender(path, type, start_page=1):
    global recommender
    if recommender is None:
        recommender = SemanticSearch()

    texts = pdf_to_text(path, start_page=start_page)[text_type[type]]
    chunks = text_to_chunks(texts)
    recommender.fit(chunks)
    return 'Corpus Loaded.'

# Query LLM

def generate_text(openAI_key,prompt, model_to_use, n_agents):
    openai.api_key = openAI_key
    messages = [{'role': 'system', 'content': f'You are a helpful assistant who identifies as a {identity} and are reading a study about {topic}.'},
                {'role': 'user', 'content': prompt}]
    completions = openai.ChatCompletion.create(
                    model=model_to_use,
                    messages=messages,
                    max_tokens=512,
                # logprobs = logprobs,
                    n=n_agents,
                    stop=None,
                    temperature=temperature)
    message = completions.choices[0].message['content']
    return message
    

def generate_answer(question,openAI_api_key):
    topn_chunks = recommender(question)
   #print(topn_chunks)
    prompt = ""
    prompt += 'Study:\n\n'
    for c in topn_chunks:
        prompt += c + '\n\n'
        
    prompt += base_prompt
    prompt += f"Instructions: Extract {prompt_type[q_num]} data from the study provided according the the Query. "
    prompt += f"Data Format: {data_format[prompt_type[q_num]][1]}"
    prompt += f"Query: {question}\nAnswer:"

    answer = generate_text(openAI_api_key, prompt, model_to_use,n_agents)
    return answer

def question_answer(url, file, question,openAI_api_key,type):
    if openAI_api_key.strip()=='':
        return '[ERROR]: Please enter you Open AI Key. Get your key here : https://platform.openai.com/account/api-keys'
    if url.strip() == '' and file == None:
        return '[ERROR]: Both URL and PDF is empty. Provide atleast one.'
    
    if url.strip() != '' and file != None:
        return '[ERROR]: Both URL and PDF is provided. Please provide only one (eiter URL or PDF).'

    if url.strip() != '':
        glob_url = url
        download_pdf(glob_url, 'corpus.pdf')
        load_recommender('corpus.pdf',type)
    
    if file != '':
        load_recommender('corpus.pdf',type)

    else:
        old_file_name = file.name
        file_name = file.name
        file_name = file_name[:-12] + file_name[-4:]
        os.rename(old_file_name, file_name)
        load_recommender(file_name,type)

    if question.strip() == '':
        return '[ERROR]: Question field is empty'

    return generate_answer(question,openAI_api_key)

def respond(question,type):
    n_attempts = 0
    while n_attempts<n_retries: # Force a correctly formatted response.
        # Add in random string for uniqueness
        if rand_seed:
            rand_string = ''.join(random.choice(string.ascii_letters + string.digits) for _ in range(20))
            new_question = f"Ignore this string: {rand_string}" + "\n\n" + question             
        else:
            new_question = question
        try:
            response = question_answer(url, file, new_question,openAI_api_key,type)#.choices[0].message['content']
            print('\n' + response + '\n')
            if 'No Relevant Data' in response:
                r_split = ["NA"]*2
            elif '%%' in response:
                r_split = response.split("%%")
            else:
                r_split = response.split("Context")
            n_attempts = n_retries + 1
        except:
                print("Bad Parsing...Trying again...")
                n_attempts += 1
    return response, r_split[0], r_split[1]

def save_sheet(identity,df,out_path):

  ### Save Data
  # Define the file path
  file_name = out_path + '.xlsx'
  print("Saving to " + file_name)
  if not os.path.exists(file_name):
      workbook = openpyxl.Workbook()
      workbook.save(file_name)
      # Open the Excel file using Pandas ExcelWriter
      excel_writer = pd.ExcelWriter(file_name, engine='openpyxl')

      # Write the DataFrame to a new sheet
      sheet_name = identity + '-' + str(1)
      df.to_excel(excel_writer, sheet_name=sheet_name, index=False)
      # Save the changes to the Excel file
      excel_writer._save()
      print("\n" + f"Saving Outputs to Sheet {sheet_name}" + "\n")
  else:
      # If the file exists, open it using openpyxl and add the DataFrame to a new sheet
      workbook = openpyxl.load_workbook(filename=file_name)
      sheet_name = identity
      counter = 0
      while sheet_name in workbook.sheetnames:
          counter += 1
          sheet_name = identity + '-' + str(counter)
      print("\n" + f"Saving Outputs to Sheet {sheet_name}" + "\n")
      excel_writer = pd.ExcelWriter(file_name, engine='openpyxl', mode='a')
      df.to_excel(excel_writer, sheet_name=sheet_name, index=False)
      excel_writer._save()
      excel_writer.close()



