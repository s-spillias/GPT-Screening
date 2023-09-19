
import re
import fitz
import numpy as np
import os
import openpyxl
import pandas as pd
import pytesseract
from pdf2image import convert_from_path
from PIL import Image
from tempfile import TemporaryDirectory
from fuzzywuzzy import fuzz,process
import openai
import time

def match_title(screen_file,paper):
    # Read in Screened spreadsheet
    df = pd.read_excel(screen_file, sheet_name=os.environ.get("SCREEN_NAME") + "_summary")
    df = df[df['Accept'] == 'Yes']

    # Assuming your titles are in a column named 'Title'
    titles = df['Title'].tolist()

    match_found = False
    proj_location = os.environ.get("PROJ_LOCATION")
    pdf_location = proj_location + '/' + os.environ.get('PDF_LOCATION')
    pdf_title = paper.replace(pdf_location,"").lower().split(" - ")[-1].replace('.pdf','')
    length = len(pdf_title)
    for title in titles:
        score = fuzz.ratio(pdf_title,title.lower()[:length])
        
        # You can adjust the threshold as needed
        if score >= 90:  # You can adjust the similarity threshold (e.g., 80)
            match_found = True
            print(title)
            break  # Exit the loop once a match is found
    return match_found

def get_abstract(truncated_title, dataframe):
    # Use fuzzy string matching to find the best match from the "Title" column
    best_match = process.extractOne(truncated_title, dataframe['Title'])

    # Check if the fuzzy score exceeds a certain threshold (e.g., 80)
    if best_match[1] >= 80:
        # Find the index of the best match
        best_match_index = dataframe[dataframe['Title'] == best_match[0]].index[0]

        # Retrieve the corresponding abstract using the index
        relevant_abstract = dataframe.loc[best_match_index, 'Abstract']
        return relevant_abstract
    else:
        # Handle cases where there's no good match
        return None

def colors(integer):
    color_permutations = [
        (0.1, 0.5, 0.8),  # Blue shade
        (0.7, 0.2, 0.1),  # Red shade
        (0.2, 0.6, 0.3),  # Green shade
        (0.8, 0.7, 0.1),  # Yellow shade
        (0.5, 0.2, 0.7),  # Purple shade
        # Add more color permutations as needed
    ]
    
    selected_color = color_permutations[integer % len(color_permutations)]
    stroke_color = selected_color
    color_dict = {
        "stroke": stroke_color,
    }
    return color_dict


def highlight_PDF(paper, phrases, paper_highlight,q_num):
   # paper_highlight = paper_highlight
    if os.path.exists(paper_highlight):
        doc = fitz.open(paper_highlight)
        incremental = True
    else:
        doc = fitz.open(paper)
        incremental = False
    
    for page in doc:
        for phrase in phrases:            
            text_instances = page.search_for(phrase)

            for inst in text_instances:
                highlight = page.add_highlight_annot(inst)
                highlight.set_colors(colors(q_num))
                highlight.update()
    if incremental:
        doc.save(paper_highlight, incremental=True, encryption=fitz.PDF_ENCRYPT_KEEP)
    else:
        doc.save(paper_highlight)

def save_row(identity,df,out_path):
    # Define the file path
    file_name = out_path + '.xlsx'
    print("Saving to " + file_name)
    if not os.path.exists(file_name):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = identity
        workbook.save(file_name)
        workbook.close()
    # Write the DataFrame to a new sheet
    workbook = openpyxl.load_workbook(filename=file_name)
    sheet_name = identity
    if sheet_name not in workbook.sheetnames:
        # If the sheet doesn't exist, create it
        sheet = workbook.create_sheet(title=sheet_name)
    else:
        sheet = workbook[sheet_name]
    print("\n" + f"Saving Outputs to Sheet {sheet_name}" + "\n")
    sheet.append(df)
    workbook.save(file_name)
    workbook.close()

def save_sheet(identity,df,out_path):

  ### Save Data
  # Define the file path
  file_name = out_path + '.xlsx'
  print("Saving to " + file_name)
  if not os.path.exists(file_name):
      workbook = openpyxl.Workbook()
      workbook.save(file_name)
      # Open the Excel file using Pandas ExcelWriter
      excel_writer = pd.ExcelWriter(file_name, engine='openpyxl')

      # Write the DataFrame to a new sheet
      sheet_name = identity + '-' + str(1)
      df.to_excel(excel_writer, sheet_name=sheet_name, index=False)
      # Save the changes to the Excel file
      excel_writer._save()
      print("\n" + f"Saving Outputs to Sheet {sheet_name}" + "\n")
  else:
      # If the file exists, open it using openpyxl and add the DataFrame to a new sheet
      workbook = openpyxl.load_workbook(filename=file_name)
      sheet_name = identity
      counter = 0
      while sheet_name in workbook.sheetnames:
          counter += 1
          sheet_name = identity + '-' + str(counter)
      print("\n" + f"Saving Outputs to Sheet {sheet_name}" + "\n")
      excel_writer = pd.ExcelWriter(file_name, engine='openpyxl', mode='a')
      df.to_excel(excel_writer, sheet_name=sheet_name, index=False)
      excel_writer._save()
      excel_writer.close()



def generate_text(openAI_key, prompt, n_agents, model_to_use):
    openai.api_key = openAI_key
    messages = [{'role': 'system', 'content': 'You are a helpful assistant.'},
                {'role': 'user', 'content': prompt}]
    failed = True
    attempt = 0
    while failed:
        attempt += 1
        if attempt < int(os.environ.get("N_RETRIES")):
            try:
                completions = openai.ChatCompletion.create(
                    model=model_to_use,
                    messages=messages,
                    max_tokens=512,
                # logprobs = logprobs,
                    n=n_agents,
                    stop=None,
                    temperature= int(os.environ.get("TEMPERATURE")))
                failed = False
            except:
                print('Connection Error - Retrying')
                time.sleep(1*2^attempt)     
        else:    
            continue
   # message = completions.choices.message['content']
    return completions